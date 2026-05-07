from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


ROOT = Path(__file__).resolve().parents[1]
TEMPLATE_PATH = ROOT / "excel_files" / "templates" / "ECC_TEMPLATE.xlsx"

SUMMARY_SHEET = "汇总"
PROD_SHEET = "生产事件单处理情况"
OS_VERSION_SHEET = "操作系统版本升级事件单处理情况"
OS_BASELINE_SHEET = "操作系统基线要求升级事件单处理情况"

PROD_HEADERS = ["工单编号", "创建人", "流程描述", "当前节点", "当前处理人/角色", "送达时间", "创建时间", "结束时间"]
OS_HEADERS = ["工单编号", "创建人", "流程描述", "当前节点", "当前处理人/角色", "送达时间", "创建时间", "结束时间", "是否启阅", "主机"]
ALERT_HEADERS = ["序号", "IP地址", "告警来源", "级别", "所属应用", "负责人", "告警信息"]


def setup_common_widths(ws, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def fill_block(ws, min_row: int, max_row: int, min_col: int, max_col: int, color: str) -> None:
    fill = PatternFill("solid", fgColor=color)
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            ws.cell(row, col).fill = fill


def border_block(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    thin = Side(style="thin", color="FF000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            ws.cell(row, col).border = border


def style_table(ws, header_row: int, first_data_row: int, last_row: int, last_col: int) -> None:
    thin = Side(style="thin", color="FF000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="FFFFFF00")
    for row in range(header_row, last_row + 1):
        for col in range(1, last_col + 1):
            cell = ws.cell(row, col)
            cell.border = border
            cell.alignment = Alignment(horizontal="center" if col == 1 else "left", vertical="center", wrap_text=True)
            if row == header_row:
                cell.fill = header_fill
                cell.font = Font(name="宋体", size=11, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif row >= first_data_row:
                cell.font = Font(name="宋体", size=11)


def build_summary_sheet(ws) -> None:
    for rng in ("A1:E1", "A6:E6", "A9:E9", "A13:E13", "A17:C17", "A21:D21", "A25:E25", "A7:E7"):
        ws.merge_cells(rng)

    ws["A1"] = "监控告警统计"
    headers = ["日期", "值班人", "监控告警数量", "邮件电话通知数量", "青藤云"]
    sample_monitor = ["示例日期", "", 0, 0, 0]
    total_monitor = ["总计", "", 0, 0, 0]
    for col, value in enumerate(headers, 1):
        ws.cell(2, col).value = value
    for col, value in enumerate(sample_monitor, 1):
        ws.cell(3, col).value = value
    for col, value in enumerate(total_monitor, 1):
        ws.cell(4, col).value = value

    ws["A6"] = "青藤云安全告警汇总"
    ws["A7"] = "示例青藤云汇总"
    ws["A9"] = "生产事件单统计"
    for col, value in enumerate(["总数", "审结", "处理中/事件录入", "超时", "审结率"], 1):
        ws.cell(10, col).value = value
    for col, value in enumerate([0, 0, 0, 0, "0.00%"], 1):
        ws.cell(11, col).value = value

    ws["A13"] = "操作系统版本升级事件单统计"
    for col, value in enumerate(["总数", "审结", "安全团队待处理", "负责人待审核", "处理中"], 1):
        ws.cell(14, col).value = value
    for col, value in enumerate([0, 0, 0, 0, 0], 1):
        ws.cell(15, col).value = value

    ws["A17"] = "告警类型统计"
    for col, value in enumerate(["事件类型", "出现次数", "占比"], 1):
        ws.cell(18, col).value = value
    for col, value in enumerate(["示例事件", 0, "0.00%"], 1):
        ws.cell(19, col).value = value

    ws["A21"] = "高频告警事件"
    for col, value in enumerate(["告警详情", "出现次数", "应用", "管理员"], 1):
        ws.cell(22, col).value = value
    for col, value in enumerate(["示例告警详情", 0, "", ""], 1):
        ws.cell(23, col).value = value

    ws["A25"] = "青藤云异常登录事件: 0"

    fill_block(ws, 1, 4, 1, 5, "FFE2F0D9")
    fill_block(ws, 6, 7, 1, 5, "FFFFC000")
    fill_block(ws, 9, 15, 1, 5, "FFD9E2F3")
    fill_block(ws, 17, 19, 1, 3, "FFFCE4D6")
    fill_block(ws, 21, 23, 1, 4, "FFFFF2CC")
    fill_block(ws, 25, 25, 1, 5, "FFE2F0D9")
    fill_block(ws, 10, 11, 4, 4, "FFFFFF00")

    for row in (1, 6, 9, 13, 17, 21, 25):
        for col in range(1, 6):
            ws.cell(row, col).font = Font(name="宋体", size=14, bold=True)
            ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in (2, 4, 10, 11, 14, 15, 18, 22):
        for col in range(1, 6):
            ws.cell(row, col).font = Font(name="宋体", size=12, bold=row in (2, 4, 10, 14, 18, 22))
    ws["A7"].font = Font(name="宋体", size=11, bold=True)
    ws["A7"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[7].height = 34
    ws["A25"].alignment = Alignment(horizontal="center", vertical="center")

    border_block(ws, 1, 4, 1, 5)
    border_block(ws, 6, 7, 1, 5)
    border_block(ws, 9, 15, 1, 5)
    border_block(ws, 17, 19, 1, 3)
    border_block(ws, 21, 23, 1, 4)
    border_block(ws, 25, 25, 1, 5)

    setup_common_widths(ws, {"A": 24, "B": 16, "C": 20, "D": 20, "E": 14, "G": 16, "H": 16, "I": 16, "J": 16, "K": 16, "L": 16})

    bar = BarChart()
    bar.style = 10
    bar.title = "监控告警统计"
    bar.y_axis.title = "数量"
    data = Reference(ws, min_col=3, max_col=5, min_row=2, max_row=4)
    cats = Reference(ws, min_col=1, min_row=3, max_row=4)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    colors = ("4472C4", "ED7D31", "FFC000")
    for idx, series in enumerate(bar.series):
        series.graphicalProperties.solidFill = colors[idx]
        series.graphicalProperties.line.solidFill = colors[idx]
    bar.height = 8.5
    bar.width = 18
    ws.add_chart(bar, "G1")

    pie = PieChart()
    pie.style = 10
    pie.title = "事件单审结率"
    data = Reference(ws, min_col=2, max_col=3, min_row=11, max_row=11)
    cats = Reference(ws, min_col=2, max_col=3, min_row=10, max_row=10)
    pie.add_data(data, from_rows=True)
    pie.set_categories(cats)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showCatName = True
    pie.dataLabels.showLeaderLines = True
    pie.firstSliceAng = 270
    pie.height = 8.5
    pie.width = 18
    ws.add_chart(pie, "G20")

    type_chart = BarChart()
    type_chart.type = "bar"
    type_chart.barDir = "bar"
    type_chart.style = 10
    type_chart.title = "告警类型统计"
    data = Reference(ws, min_col=2, min_row=18, max_row=19)
    cats = Reference(ws, min_col=1, min_row=19, max_row=19)
    type_chart.add_data(data, titles_from_data=True)
    type_chart.set_categories(cats)
    type_chart.legend = None
    type_chart.gapWidth = 45
    type_chart.x_axis.scaling.min = 0
    type_chart.x_axis.scaling.max = 5
    for series in type_chart.series:
        series.graphicalProperties.solidFill = "ED7D31"
        series.graphicalProperties.line.solidFill = "ED7D31"
        series.dLbls = DataLabelList()
        series.dLbls.showVal = True
    type_chart.height = 5.8
    type_chart.width = 13.5
    ws.add_chart(type_chart, "G39")


def build_production_sheet(ws) -> None:
    ws.merge_cells("A1:H1")
    ws["A1"] = "生产事件工单"
    ws["A1"].font = Font(name="宋体", size=16, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    for col, header in enumerate(PROD_HEADERS, 1):
        ws.cell(2, col).value = header
    style_table(ws, 2, 3, 3, len(PROD_HEADERS))
    for col in range(1, len(PROD_HEADERS) + 1):
        ws.cell(3, col).fill = PatternFill("solid", fgColor="FFFFFFFF")
        ws.cell(3, col).font = Font(name="宋体", size=11)
    setup_common_widths(ws, {"A": 18, "B": 12, "C": 80, "D": 16, "E": 22, "F": 22, "G": 22, "H": 22})


def build_os_version_sheet(ws) -> None:
    for col, header in enumerate(OS_HEADERS, 1):
        ws.cell(1, col).value = header
    style_table(ws, 1, 2, 2, len(OS_HEADERS))
    setup_common_widths(ws, {"A": 22, "B": 12, "C": 60, "D": 24, "E": 22, "F": 22, "G": 22, "H": 22, "I": 12, "J": 18})


def build_os_baseline_sheet(ws) -> None:
    for col, header in enumerate(OS_HEADERS, 1):
        ws.cell(1, col).value = header
    style_table(ws, 1, 2, 2, len(OS_HEADERS))
    for col in range(1, len(OS_HEADERS) + 1):
        ws.cell(2, col).fill = PatternFill("solid", fgColor="FFFFFFFF")
        ws.cell(2, col).font = Font(name="宋体", size=11)
    setup_common_widths(ws, {"A": 22, "B": 12, "C": 60, "D": 24, "E": 22, "F": 22, "G": 22, "H": 22, "I": 12, "J": 18})


def build_template() -> Path:
    TEMPLATE_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.active.title = SUMMARY_SHEET
    wb.create_sheet(PROD_SHEET)
    wb.create_sheet(OS_VERSION_SHEET)
    wb.create_sheet(OS_BASELINE_SHEET)

    build_summary_sheet(wb[SUMMARY_SHEET])
    build_production_sheet(wb[PROD_SHEET])
    build_os_version_sheet(wb[OS_VERSION_SHEET])
    build_os_baseline_sheet(wb[OS_BASELINE_SHEET])

    wb.save(TEMPLATE_PATH)
    return TEMPLATE_PATH


if __name__ == "__main__":
    path = build_template()
    print(path)
