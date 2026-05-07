import logging
import os
import json
from typing import Any, List, Dict, Optional

from mcp.server.fastmcp import FastMCP
from mcp.types import ToolAnnotations

# Import exceptions
from excel_mcp.exceptions import (
    ValidationError,
    WorkbookError,
    SheetError,
    DataError,
    FormattingError,
    CalculationError,
    PivotError,
    ChartError
)

# Import from excel_mcp package with consistent _impl suffixes
from excel_mcp.validation import (
    validate_formula_in_cell_operation as validate_formula_impl,
    validate_range_in_sheet_operation as validate_range_impl
)
from excel_mcp.chart import create_chart_in_sheet as create_chart_impl
from excel_mcp.workbook import get_workbook_info
from excel_mcp.data import write_data
from excel_mcp.pivot import create_pivot_table as create_pivot_table_impl
from excel_mcp.tables import create_excel_table as create_table_impl
from excel_mcp.ecc_report import (
    generate_ecc_report_from_template as generate_ecc_report_from_template_impl,
    generate_ecc_report_from_template_files as generate_ecc_report_from_template_files_impl,
)
from excel_mcp.sheet import (
    copy_sheet,
    delete_sheet,
    rename_sheet,
    merge_range,
    unmerge_range,
    get_merged_ranges,
    insert_row,
    insert_cols,
    delete_rows,
    delete_cols,
)

# Get project root directory path for log file path.
# When using the stdio transmission method,
# relative paths may cause log files to fail to create
# due to the client's running location and permission issues,
# resulting in the program not being able to run.
# Thus using os.path.join(ROOT_DIR, "excel-mcp.log") instead.

ROOT_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
LOG_FILE = os.path.join(ROOT_DIR, "excel-mcp.log")

# Initialize EXCEL_FILES_PATH variable without assigning a value
EXCEL_FILES_PATH = None

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    handlers=[
        # Referring to https://github.com/modelcontextprotocol/python-sdk/issues/409#issuecomment-2816831318
        # The stdio mode server MUST NOT write anything to its stdout that is not a valid MCP message.
        logging.FileHandler(LOG_FILE)
    ],
)
logger = logging.getLogger("excel-mcp")
# Initialize FastMCP server
mcp = FastMCP(
    "excel-mcp",
    host=os.environ.get("FASTMCP_HOST", "0.0.0.0"),
    port=int(os.environ.get("FASTMCP_PORT", "8017")),
    instructions="Excel MCP Server for manipulating Excel files"
)


def _resolved_path_is_within(base: str, candidate: str) -> bool:
    base = os.path.realpath(base)
    candidate = os.path.realpath(candidate)
    if candidate == base:
        return True
    try:
        return os.path.commonpath([base, candidate]) == base
    except ValueError:
        return False


def get_excel_path(filename: str) -> str:
    """Get full path to Excel file.

    Args:
        filename: Name of Excel file

    Returns:
        Full path to Excel file
    """
    if not filename or "\x00" in filename:
        raise ValueError(f"Invalid filename: {filename}")

    if EXCEL_FILES_PATH is None:
        if not os.path.isabs(filename):
            raise ValueError(f"Invalid filename: {filename}, must be an absolute path when not in SSE mode")
        return os.path.normpath(filename)

    if os.path.isabs(filename):
        raise ValueError(f"Invalid filename: {filename}, must be relative to EXCEL_FILES_PATH")

    base = os.path.realpath(EXCEL_FILES_PATH)
    candidate = os.path.realpath(os.path.join(base, filename))

    if not _resolved_path_is_within(base, candidate):
        raise ValueError(f"Invalid filename: {filename}, path escapes EXCEL_FILES_PATH")

    return candidate

@mcp.tool(
    annotations=ToolAnnotations(
        title="Apply Formula",
        destructiveHint=True,
    ),
)
def apply_formula(
    filepath: str,
    sheet_name: str,
    cell: str,
    formula: str,
) -> str:
    """
    Apply Excel formula to cell.
    Excel formula will write to cell with verification.
    """
    try:
        full_path = get_excel_path(filepath)
        # First validate the formula
        validation = validate_formula_impl(full_path, sheet_name, cell, formula)
        if isinstance(validation, dict) and "error" in validation:
            return f"Error: {validation['error']}"
            
        # If valid, apply the formula
        from excel_mcp.calculations import apply_formula as apply_formula_impl
        result = apply_formula_impl(full_path, sheet_name, cell, formula)
        return result["message"]
    except (ValidationError, CalculationError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error applying formula: {e}")
        raise

@mcp.tool(
    annotations=ToolAnnotations(
        title="Validate Formula Syntax",
        readOnlyHint=True,
    ),
)
def validate_formula_syntax(
    filepath: str,
    sheet_name: str,
    cell: str,
    formula: str,
) -> str:
    """Validate Excel formula syntax without applying it."""
    try:
        full_path = get_excel_path(filepath)
        result = validate_formula_impl(full_path, sheet_name, cell, formula)
        return result["message"]
    except (ValidationError, CalculationError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error validating formula: {e}")
        raise

@mcp.tool(
    annotations=ToolAnnotations(
        title="Format Range",
        destructiveHint=True,
    ),
)
def format_range(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    end_cell: Optional[str] = None,
    bold: bool = False,
    italic: bool = False,
    underline: bool = False,
    font_size: Optional[int] = None,
    font_color: Optional[str] = None,
    bg_color: Optional[str] = None,
    border_style: Optional[str] = None,
    border_color: Optional[str] = None,
    number_format: Optional[str] = None,
    alignment: Optional[str] = None,
    wrap_text: bool = False,
    merge_cells: bool = False,
    protection: Optional[Dict[str, Any]] = None,
    conditional_format: Optional[Dict[str, Any]] = None
) -> str:
    """Apply formatting to a range of cells."""
    try:
        full_path = get_excel_path(filepath)
        from excel_mcp.formatting import format_range as format_range_func
        
        # Convert None values to appropriate defaults for the underlying function
        format_range_func(
            filepath=full_path,
            sheet_name=sheet_name,
            start_cell=start_cell,
            end_cell=end_cell,  # This can be None
            bold=bold,
            italic=italic,
            underline=underline,
            font_size=font_size,  # This can be None
            font_color=font_color,  # This can be None
            bg_color=bg_color,  # This can be None
            border_style=border_style,  # This can be None
            border_color=border_color,  # This can be None
            number_format=number_format,  # This can be None
            alignment=alignment,  # This can be None
            wrap_text=wrap_text,
            merge_cells=merge_cells,
            protection=protection,  # This can be None
            conditional_format=conditional_format  # This can be None
        )
        return "Range formatted successfully"
    except (ValidationError, FormattingError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error formatting range: {e}")
        raise

@mcp.tool(
    annotations=ToolAnnotations(
        title="Read Data from Excel",
        readOnlyHint=True,
    ),
)
def read_data_from_excel(
    filepath: str,
    sheet_name: str,
    start_cell: str = "A1",
    end_cell: Optional[str] = None,
    preview_only: bool = False
) -> str:
    """
    Read data from Excel worksheet with cell metadata including validation rules.
    
    Args:
        filepath: Path to Excel file
        sheet_name: Name of worksheet
        start_cell: Starting cell (default A1)
        end_cell: Ending cell (optional, auto-expands if not provided)
        preview_only: Whether to return preview only
    
    Returns:  
    JSON string containing structured cell data with validation metadata.
    Each cell includes: address, value, row, column, and validation info (if any).
    """
    try:
        full_path = get_excel_path(filepath)
        from excel_mcp.data import read_excel_range_with_metadata
        result = read_excel_range_with_metadata(
            full_path, 
            sheet_name, 
            start_cell, 
            end_cell
        )
        if not result or not result.get("cells"):
            return "No data found in specified range"
            
        # Return as formatted JSON string
        import json
        return json.dumps(result, indent=2, default=str)
        
    except Exception as e:
        logger.error(f"Error reading data: {e}")
        raise

@mcp.tool(
    annotations=ToolAnnotations(
        title="Write Data to Excel",
        destructiveHint=True,
    ),
)
def write_data_to_excel(
    filepath: str,
    sheet_name: str,
    data: List[List],
    start_cell: str = "A1",
) -> str:
    """
    Write data to Excel worksheet.
    Excel formula will write to cell without any verification.

    PARAMETERS:  
    filepath: Path to Excel file
    sheet_name: Name of worksheet to write to
    data: List of lists containing data to write to the worksheet, sublists are assumed to be rows
    start_cell: Cell to start writing to, default is "A1"
  
    """
    try:
        full_path = get_excel_path(filepath)
        result = write_data(full_path, sheet_name, data, start_cell)
        return result["message"]
    except (ValidationError, DataError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error writing data: {e}")
        raise

@mcp.tool(
    annotations=ToolAnnotations(
        title="Create Workbook",
        destructiveHint=True,
    ),
)
def create_workbook(filepath: str) -> str:
    """Create new Excel workbook."""
    try:
        full_path = get_excel_path(filepath)
        from excel_mcp.workbook import create_workbook as create_workbook_impl
        create_workbook_impl(full_path)
        return f"Created workbook at {full_path}"
    except WorkbookError as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating workbook: {e}")
        raise

@mcp.tool(
    annotations=ToolAnnotations(
        title="Generate ECC Report from Template",
        destructiveHint=True,
    ),
)
def generate_ecc_report_from_template(
    production_events: Any,
    os_baseline_events: Any,
    omp_alerts: Any,
    qingteng_events: Any,
    qingteng_summary: Any,
    qingteng_abnormal_login_count: Any,
    output_file: str,
) -> str:
    """Generate ECC report xlsx from cleaned JSON data using the fixed template."""
    try:
        full_output = get_excel_path(output_file)
        result = generate_ecc_report_from_template_impl(
            template_file=get_excel_path("templates/ECC_TEMPLATE.xlsx"),
            output_file=full_output,
            production_events=production_events,
            os_baseline_events=os_baseline_events,
            omp_alerts=omp_alerts,
            qingteng_events=qingteng_events,
            qingteng_summary=qingteng_summary,
            qingteng_abnormal_login_count=qingteng_abnormal_login_count,
        )
        return json.dumps(result, ensure_ascii=False, indent=2)
    except (WorkbookError, DataError, ValidationError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error generating ECC report from template: {e}")
        raise

@mcp.tool(
    annotations=ToolAnnotations(
        title="Generate ECC Report from Template Files",
        destructiveHint=True,
    ),
)
def generate_ecc_report_from_template_files(
    production_events_file: str,
    os_baseline_events_file: str,
    omp_alerts_file: str,
    qingteng_events_file: str,
    qingteng_summary_file: str,
    qingteng_abnormal_login_count_file: str,
    output_file: str,
) -> str:
    """Generate ECC report xlsx from JSON files using the fixed template."""
    try:
        full_output = get_excel_path(output_file)
        result = generate_ecc_report_from_template_files_impl(
            template_file=get_excel_path("templates/ECC_TEMPLATE.xlsx"),
            output_file=full_output,
            production_events_file=get_excel_path(production_events_file),
            os_baseline_events_file=get_excel_path(os_baseline_events_file),
            omp_alerts_file=get_excel_path(omp_alerts_file),
            qingteng_events_file=get_excel_path(qingteng_events_file),
            qingteng_summary_file=get_excel_path(qingteng_summary_file),
            qingteng_abnormal_login_count_file=get_excel_path(qingteng_abnormal_login_count_file),
        )
        return json.dumps(result, ensure_ascii=False, indent=2)
    except (WorkbookError, DataError, ValidationError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error generating ECC report from template files: {e}")
        raise

@mcp.tool(
    annotations=ToolAnnotations(
        title="Create Worksheet",
        destructiveHint=True,
    ),
)
def create_worksheet(filepath: str, sheet_name: str) -> str:
    """Create new worksheet in workbook."""
    try:
        full_path = get_excel_path(filepath)
        from excel_mcp.workbook import create_sheet as create_worksheet_impl
        result = create_worksheet_impl(full_path, sheet_name)
        return result["message"]
    except (ValidationError, WorkbookError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating worksheet: {e}")
        raise

@mcp.tool(
    annotations=ToolAnnotations(
        title="Create Chart",
        destructiveHint=True,
    ),
)
def create_chart(
    filepath: str,
    sheet_name: str,
    data_range: str,
    chart_type: str,
    target_cell: str,
    title: str = "",
    x_axis: str = "",
    y_axis: str = ""
) -> str:
    """Create chart in worksheet."""
    try:
        full_path = get_excel_path(filepath)
        result = create_chart_impl(
            filepath=full_path,
            sheet_name=sheet_name,
            data_range=data_range,
            chart_type=chart_type,
            target_cell=target_cell,
            title=title,
            x_axis=x_axis,
            y_axis=y_axis
        )
        return result["message"]
    except (ValidationError, ChartError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating chart: {e}")
        raise

@mcp.tool(
    annotations=ToolAnnotations(
        title="Create Pivot Table",
        destructiveHint=True,
    ),
)
def create_pivot_table(
    filepath: str,
    sheet_name: str,
    data_range: str,
    rows: List[str],
    values: List[str],
    columns: Optional[List[str]] = None,
    agg_func: str = "mean"
) -> str:
    """Create pivot table in worksheet."""
    try:
        full_path = get_excel_path(filepath)
        result = create_pivot_table_impl(
            filepath=full_path,
            sheet_name=sheet_name,
            data_range=data_range,
            rows=rows,
            values=values,
            columns=columns or [],
            agg_func=agg_func
        )
        return result["message"]
    except (ValidationError, PivotError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating pivot table: {e}")
        raise

@mcp.tool(
    annotations=ToolAnnotations(
        title="Create Table",
        destructiveHint=True,
    ),
)
def create_table(
    filepath: str,
    sheet_name: str,
    data_range: str,
    table_name: Optional[str] = None,
    table_style: str = "TableStyleMedium9"
) -> str:
    """Creates a native Excel table from a specified range of data."""
    try:
        full_path = get_excel_path(filepath)
        result = create_table_impl(
            filepath=full_path,
            sheet_name=sheet_name,
            data_range=data_range,
            table_name=table_name,
            table_style=table_style
        )
        return result["message"]
    except DataError as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating table: {e}")
        raise

@mcp.tool(
    annotations=ToolAnnotations(
        title="Copy Worksheet",
        destructiveHint=True,
    ),
)
def copy_worksheet(
    filepath: str,
    source_sheet: str,
    target_sheet: str
) -> str:
    """Copy worksheet within workbook."""
    try:
        full_path = get_excel_path(filepath)
        result = copy_sheet(full_path, source_sheet, target_sheet)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error copying worksheet: {e}")
        raise

@mcp.tool(
    annotations=ToolAnnotations(
        title="Delete Worksheet",
        destructiveHint=True,
    ),
)
def delete_worksheet(
    filepath: str,
    sheet_name: str
) -> str:
    """Delete worksheet from workbook."""
    try:
        full_path = get_excel_path(filepath)
        result = delete_sheet(full_path, sheet_name)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error deleting worksheet: {e}")
        raise

@mcp.tool(
    annotations=ToolAnnotations(
        title="Rename Worksheet",
        destructiveHint=True,
    ),
)
def rename_worksheet(
    filepath: str,
    old_name: str,
    new_name: str
) -> str:
    """Rename worksheet in workbook."""
    try:
        full_path = get_excel_path(filepath)
        result = rename_sheet(full_path, old_name, new_name)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error renaming worksheet: {e}")
        raise

@mcp.tool(
    annotations=ToolAnnotations(
        title="Get Workbook Metadata",
        readOnlyHint=True,
    ),
)
def get_workbook_metadata(
    filepath: str,
    include_ranges: bool = False
) -> str:
    """Get metadata about workbook including sheets, ranges, etc."""
    try:
        full_path = get_excel_path(filepath)
        result = get_workbook_info(full_path, include_ranges=include_ranges)
        return str(result)
    except WorkbookError as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error getting workbook metadata: {e}")
        raise

@mcp.tool(
    annotations=ToolAnnotations(
        title="Merge Cells",
        destructiveHint=True,
    ),
)
def merge_cells(filepath: str, sheet_name: str, start_cell: str, end_cell: str) -> str:
    """Merge a range of cells."""
    try:
        full_path = get_excel_path(filepath)
        result = merge_range(full_path, sheet_name, start_cell, end_cell)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error merging cells: {e}")
        raise

@mcp.tool(
    annotations=ToolAnnotations(
        title="Unmerge Cells",
        destructiveHint=True,
    ),
)
def unmerge_cells(filepath: str, sheet_name: str, start_cell: str, end_cell: str) -> str:
    """Unmerge a range of cells."""
    try:
        full_path = get_excel_path(filepath)
        result = unmerge_range(full_path, sheet_name, start_cell, end_cell)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error unmerging cells: {e}")
        raise

@mcp.tool(
    annotations=ToolAnnotations(
        title="Get Merged Cells",
        readOnlyHint=True,
    ),
)
def get_merged_cells(filepath: str, sheet_name: str) -> str:
    """Get merged cells in a worksheet."""
    try:
        full_path = get_excel_path(filepath)
        return str(get_merged_ranges(full_path, sheet_name))
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error getting merged cells: {e}")
        raise

@mcp.tool(
    annotations=ToolAnnotations(
        title="Copy Range",
        destructiveHint=True,
    ),
)
def copy_range(
    filepath: str,
    sheet_name: str,
    source_start: str,
    source_end: str,
    target_start: str,
    target_sheet: Optional[str] = None
) -> str:
    """Copy a range of cells to another location."""
    try:
        full_path = get_excel_path(filepath)
        from excel_mcp.sheet import copy_range_operation
        result = copy_range_operation(
            full_path,
            sheet_name,
            source_start,
            source_end,
            target_start,
            target_sheet or sheet_name  # Use source sheet if target_sheet is None
        )
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error copying range: {e}")
        raise

@mcp.tool(
    annotations=ToolAnnotations(
        title="Delete Range",
        destructiveHint=True,
    ),
)
def delete_range(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    end_cell: str,
    shift_direction: str = "up"
) -> str:
    """Delete a range of cells and shift remaining cells."""
    try:
        full_path = get_excel_path(filepath)
        from excel_mcp.sheet import delete_range_operation
        result = delete_range_operation(
            full_path,
            sheet_name,
            start_cell,
            end_cell,
            shift_direction
        )
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error deleting range: {e}")
        raise

@mcp.tool(
    annotations=ToolAnnotations(
        title="Validate Excel Range",
        readOnlyHint=True,
    ),
)
def validate_excel_range(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    end_cell: Optional[str] = None
) -> str:
    """Validate if a range exists and is properly formatted."""
    try:
        full_path = get_excel_path(filepath)
        range_str = start_cell if not end_cell else f"{start_cell}:{end_cell}"
        result = validate_range_impl(full_path, sheet_name, range_str)
        return result["message"]
    except ValidationError as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error validating range: {e}")
        raise

@mcp.tool(
    annotations=ToolAnnotations(
        title="Get Data Validation Info",
        readOnlyHint=True,
    ),
)
def get_data_validation_info(
    filepath: str,
    sheet_name: str
) -> str:
    """
    Get all data validation rules in a worksheet.
    
    This tool helps identify which cell ranges have validation rules
    and what types of validation are applied.
    
    Args:
        filepath: Path to Excel file
        sheet_name: Name of worksheet
        
    Returns:
        JSON string containing all validation rules in the worksheet
    """
    try:
        full_path = get_excel_path(filepath)
        from openpyxl import load_workbook
        from excel_mcp.cell_validation import get_all_validation_ranges
        
        wb = load_workbook(full_path, read_only=False)
        if sheet_name not in wb.sheetnames:
            return f"Error: Sheet '{sheet_name}' not found"
            
        ws = wb[sheet_name]
        validations = get_all_validation_ranges(ws)
        wb.close()
        
        if not validations:
            return "No data validation rules found in this worksheet"
            
        import json
        return json.dumps({
            "sheet_name": sheet_name,
            "validation_rules": validations
        }, indent=2, default=str)
        
    except Exception as e:
        logger.error(f"Error getting validation info: {e}")
        raise

@mcp.tool(
    annotations=ToolAnnotations(
        title="Insert Rows",
        destructiveHint=True,
    ),
)
def insert_rows(
    filepath: str,
    sheet_name: str,
    start_row: int,
    count: int = 1
) -> str:
    """Insert one or more rows starting at the specified row."""
    try:
        full_path = get_excel_path(filepath)
        result = insert_row(full_path, sheet_name, start_row, count)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error inserting rows: {e}")
        raise

@mcp.tool(
    annotations=ToolAnnotations(
        title="Insert Columns",
        destructiveHint=True,
    ),
)
def insert_columns(
    filepath: str,
    sheet_name: str,
    start_col: int,
    count: int = 1
) -> str:
    """Insert one or more columns starting at the specified column."""
    try:
        full_path = get_excel_path(filepath)
        result = insert_cols(full_path, sheet_name, start_col, count)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error inserting columns: {e}")
        raise

@mcp.tool(
    annotations=ToolAnnotations(
        title="Delete Rows",
        destructiveHint=True,
    ),
)
def delete_sheet_rows(
    filepath: str,
    sheet_name: str,
    start_row: int,
    count: int = 1
) -> str:
    """Delete one or more rows starting at the specified row."""
    try:
        full_path = get_excel_path(filepath)
        result = delete_rows(full_path, sheet_name, start_row, count)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error deleting rows: {e}")
        raise

@mcp.tool(
    annotations=ToolAnnotations(
        title="Delete Columns",
        destructiveHint=True,
    ),
)
def delete_sheet_columns(
    filepath: str,
    sheet_name: str,
    start_col: int,
    count: int = 1
) -> str:
    """Delete one or more columns starting at the specified column."""
    try:
        full_path = get_excel_path(filepath)
        result = delete_cols(full_path, sheet_name, start_col, count)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error deleting columns: {e}")
        raise

def run_sse():
    """Run Excel MCP server in SSE mode."""
    # Assign value to EXCEL_FILES_PATH in SSE mode
    global EXCEL_FILES_PATH
    EXCEL_FILES_PATH = os.environ.get("EXCEL_FILES_PATH", "./excel_files")
    # Create directory if it doesn't exist
    os.makedirs(EXCEL_FILES_PATH, exist_ok=True)
    
    try:
        logger.info(f"Starting Excel MCP server with SSE transport (files directory: {EXCEL_FILES_PATH})")
        mcp.run(transport="sse")
    except KeyboardInterrupt:
        logger.info("Server stopped by user")
    except Exception as e:
        logger.error(f"Server failed: {e}")
        raise
    finally:
        logger.info("Server shutdown complete")

def run_streamable_http():
    """Run Excel MCP server in streamable HTTP mode."""
    # Assign value to EXCEL_FILES_PATH in streamable HTTP mode
    global EXCEL_FILES_PATH
    EXCEL_FILES_PATH = os.environ.get("EXCEL_FILES_PATH", "./excel_files")
    # Create directory if it doesn't exist
    os.makedirs(EXCEL_FILES_PATH, exist_ok=True)
    
    try:
        logger.info(f"Starting Excel MCP server with streamable HTTP transport (files directory: {EXCEL_FILES_PATH})")
        mcp.run(transport="streamable-http")
    except KeyboardInterrupt:
        logger.info("Server stopped by user")
    except Exception as e:
        logger.error(f"Server failed: {e}")
        raise
    finally:
        logger.info("Server shutdown complete")

def run_stdio():
    """Run Excel MCP server in stdio mode."""
    # No need to assign EXCEL_FILES_PATH in stdio mode
    
    try:
        logger.info("Starting Excel MCP server with stdio transport")
        mcp.run(transport="stdio")
    except KeyboardInterrupt:
        logger.info("Server stopped by user")
    except Exception as e:
        logger.error(f"Server failed: {e}")
        raise
    finally:
        logger.info("Server shutdown complete")
