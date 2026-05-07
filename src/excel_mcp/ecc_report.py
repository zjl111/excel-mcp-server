from __future__ import annotations

import ast
import json
import re
import shutil
from collections import Counter
from copy import copy
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .exceptions import DataError, WorkbookError

SUMMARY_SHEET = "汇总"
PROD_SHEET = "生产事件单处理情况"
OS_VERSION_SHEET = "操作系统版本升级事件单处理情况"
OS_BASELINE_SHEET = "操作系统基线要求升级事件单处理情况"

DEFAULT_TEMPLATE = "templates/ECC_TEMPLATE.xlsx"

PROD_HEADERS = ["工单编号", "创建人", "流程描述", "当前节点", "当前处理人/角色", "送达时间", "创建时间", "结束时间"]
OS_HEADERS = ["工单编号", "创建人", "流程描述", "当前节点", "当前处理人/角色", "送达时间", "创建时间", "结束时间", "是否启阅", "主机"]
ALERT_HEADERS = ["序号", "IP地址", "告警来源", "级别", "所属应用", "负责人", "告警信息"]


def normalize_text(text: str) -> str:
    return (
        text.replace("\ufeff", "")
        .replace("“", '"')
        .replace("”", '"')
        .replace("‘", "'")
        .replace("’", "'")
    )


def normalize_jsonish(text: str) -> str:
    text = normalize_text(text)
    result: list[str] = []
    in_string = False
    escaped = False
    for char in text:
        if char == "\\" and in_string:
            escaped = not escaped
            result.append(char)
            continue
        if char == '"' and not escaped:
            in_string = not in_string
        escaped = False
        if not in_string and char == "：":
            result.append(":")
        elif not in_string and char == "，":
            result.append(",")
        else:
            result.append(char)
    return "".join(result)


def parse_literal_block(block: Any, default: Any) -> Any:
    if isinstance(block, (list, dict)):
        return block
    if block is None or block == "":
        return default

    text = normalize_jsonish(str(block)).strip()
    if not text:
        return default

    if text.startswith("["):
        start = text.find("[")
        end = text.rfind("]")
        if start >= 0 and end >= start:
            text = text[start : end + 1]

    text = re.sub(r":\s*,", ": null,", text)
    text = re.sub(r":\s*([}\]])", r": null\1", text)
    text = re.sub(r'(?<=["0-9}\]])\s*\n\s*(?=")', ",\n", text)
    text = re.sub(r'(?<=")\s+(?=")', ", ", text)

    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pythonish = text.replace("null", "None").replace("true", "True").replace("false", "False")
        try:
            return ast.literal_eval(pythonish)
        except Exception:
            return default


def parse_int_from_text(block: Any) -> int:
    if isinstance(block, int):
        return block
    match = re.search(r"(-?\d+)", normalize_text(str(block or "")))
    return int(match.group(1)) if match else 0


def load_json_file(path: str) -> Any:
    file_path = Path(path)
    if not file_path.exists():
        raise DataError(f"JSON file not found: {file_path}")
    try:
        return json.loads(file_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise DataError(f"Invalid JSON in file {file_path}: {exc}") from exc


def make_source_data(
    production_events: Any = "",
    os_baseline_events: Any = "",
    omp_alerts: Any = "",
    qingteng_events: Any = "",
    qingteng_summary: Any = "",
    qingteng_abnormal_login_count: Any = "",
) -> dict[str, Any]:
    return {
        "production_events": list(parse_literal_block(production_events, [])),
        "os_baseline_events": list(parse_literal_block(os_baseline_events, [])),
        "omp_alerts": list(parse_literal_block(omp_alerts, [])),
        "qingteng_events": list(parse_literal_block(qingteng_events, [])),
        "qingteng_summary": list(parse_literal_block(qingteng_summary, [])),
        "qingteng_abnormal_login_count": parse_int_from_text(qingteng_abnormal_login_count),
    }


def parse_dt(value: Any) -> datetime | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, time.min)
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    return None


def parse_date_arg(value: str | None) -> date | None:
    return datetime.strptime(value, "%Y-%m-%d").date() if value else None


def parse_date_from_text(value: str) -> date | None:
    match = re.search(r"(20\d{2})[-/](\d{1,2})[-/](\d{1,2})", value or "")
    if not match:
        return None
    year, month, day = (int(part) for part in match.groups())
    return date(year, month, day)


def fmt_dt(value: Any) -> str:
    dt = parse_dt(value)
    return dt.strftime("%Y-%m-%d %H:%M:%S") if dt else ("" if value is None else str(value))


def row_value(row: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        if key in row and row[key] not in (None, ""):
            return row[key]
    return ""


def shift_date(dt: datetime) -> date:
    return (dt - timedelta(hours=8, minutes=30)).date() if dt.time() < time(8, 30) else dt.date()


def infer_range(data: dict[str, Any]) -> tuple[date, date]:
    dates: list[date] = []
    for row in data["omp_alerts"]:
        dt = parse_dt(row_value(row, "通知时间"))
        if dt:
            dates.append(shift_date(dt))
    for text in data["qingteng_summary"]:
        text_date = parse_date_from_text(str(text))
        if text_date:
            dates.append(text_date)
    if not dates:
        for row in data["qingteng_events"]:
            dt = parse_dt(row_value(row, "通知时间"))
            if dt:
                dates.append(shift_date(dt))
    if not dates:
        today = date.today()
        return today, today
    return min(dates), max(dates)


def iter_days(start: date, end: date) -> list[date]:
    return [start + timedelta(days=i) for i in range((end - start).days + 1)]


def in_day_window(dt: datetime | None, day: date) -> bool:
    if not dt:
        return False
    start = datetime.combine(day, time(8, 30))
    return start <= dt < start + timedelta(days=1)


def is_finished(row: dict[str, Any]) -> bool:
    return bool(parse_dt(row_value(row, "结束时间"))) or row_value(row, "当前节点") == "审结"


def production_report_end(end: date) -> datetime:
    return datetime.combine(end + timedelta(days=1), time(8, 30))


def is_production_timeout(row: dict[str, Any], report_end: datetime) -> bool:
    create_dt = parse_dt(row_value(row, "创建时间"))
    return bool(create_dt and not is_finished(row) and report_end - create_dt > timedelta(days=7))


def filter_production_events(rows: list[dict[str, Any]], start: date, end: date) -> list[dict[str, Any]]:
    end_dt = production_report_end(end)
    start_dt = datetime.combine(start, time(8, 30))
    result: list[dict[str, Any]] = []
    for row in rows:
        create_dt = parse_dt(row_value(row, "创建时间"))
        end_time = parse_dt(row_value(row, "结束时间"))
        if end_time:
            if start_dt <= end_time < end_dt:
                result.append(row)
        elif create_dt and create_dt < end_dt:
            result.append(row)

    def sort_key(row: dict[str, Any]) -> tuple[int, datetime]:
        create_dt = parse_dt(row_value(row, "创建时间")) or datetime.min
        timeout = is_production_timeout(row, end_dt)
        bucket = 0 if timeout else 2 if is_finished(row) else 1
        return bucket, create_dt

    return sorted(result, key=sort_key)


def sort_os_baseline_events(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return sorted(rows, key=lambda row: (1 if row_value(row, "当前节点") == "审结" else 0, parse_dt(row_value(row, "创建时间")) or datetime.min))


def alert_row(alert: dict[str, Any], source: str) -> dict[str, Any]:
    alert_source = str(row_value(alert, "告警来源") or source).replace(" ", "")
    return {
        "IP地址": row_value(alert, "IP地址", "IP 地址", "ip", "IP"),
        "告警来源": "OMP平台" if "OMP" in alert_source.upper() else "青藤云" if "青藤云" in alert_source else alert_source,
        "级别": row_value(alert, "级别"),
        "所属应用": row_value(alert, "所属应用"),
        "负责人": row_value(alert, "负责人"),
        "告警信息": row_value(alert, "告警信息"),
        "通知时间": row_value(alert, "通知时间"),
        "通知方式": row_value(alert, "通知方式"),
        "告警指标名称": row_value(alert, "告警指标名称"),
        "事件名称": row_value(alert, "事件名称"),
    }


def qingteng_summary_rows(data: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    event_times = [parse_dt(row_value(row, "通知时间")) for row in data["qingteng_events"]]
    for index, text in enumerate(data["qingteng_summary"]):
        content = str(text)
        text_date = parse_date_from_text(content)
        notify_time = None
        if text_date:
            notify_time = datetime.combine(text_date, time(20, 52, 0))
        elif index < len(event_times) and event_times[index]:
            notify_time = event_times[index]

        level_match = re.search(r"\[青藤云安全\]\[([^\]]+)\]", content)
        host_match = re.search(r"(?:/|主机\s*)(\d{1,3}(?:\.\d{1,3}){3})", content)
        event_match = re.search(r"\]\s*([^:\]]+):", content)
        rows.append(
            {
                "IP地址": host_match.group(1) if host_match else "",
                "告警来源": "青藤云",
                "级别": level_match.group(1) if level_match else "",
                "所属应用": "",
                "负责人": "",
                "告警信息": content,
                "通知时间": notify_time.strftime("%Y-%m-%d %H:%M:%S") if notify_time else "",
                "通知方式": "",
                "告警指标名称": "",
                "事件名称": event_match.group(1).strip() if event_match else "青藤云安全事件",
            }
        )
    if rows:
        return rows
    return [alert_row(row, "青藤云") for row in data["qingteng_events"]]


def all_alerts(data: dict[str, Any]) -> list[dict[str, Any]]:
    alerts = [alert_row(row, "OMP平台") for row in data["omp_alerts"]]
    alerts += qingteng_summary_rows(data)
    return alerts


def alert_days(data: dict[str, Any]) -> list[date]:
    return sorted(
        {
            shift_date(dt)
            for alert in all_alerts(data)
            for dt in [parse_dt(alert.get("通知时间"))]
            if dt
        }
    )


def percent(part: int, total: int) -> str:
    return f"{part / total * 100:.2f}%" if total else "0.00%"


def copy_cell_style(source, target) -> None:
    if source.has_style:
        target._style = copy(source._style)
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)
    target.number_format = source.number_format
    target.protection = copy(source.protection)


def clear_values(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.value = None


def write_row(ws, row_index: int, values: list[Any], style_row: int | None = None) -> None:
    if style_row:
        for col in range(1, len(values) + 1):
            if isinstance(ws.cell(row_index, col), MergedCell):
                continue
            copy_cell_style(ws.cell(style_row, col), ws.cell(row_index, col))
    for col, value in enumerate(values, 1):
        cell = ws.cell(row_index, col)
        if not isinstance(cell, MergedCell):
            cell.value = value


def reset_sheet_rows(ws, keep_rows: int = 0) -> None:
    if ws.max_row > keep_rows:
        ws.delete_rows(keep_rows + 1, ws.max_row - keep_rows)


def ensure_summary_charts(ws) -> None:
    ws._charts = []

    def nice_axis_max(values: list[int]) -> int:
        max_value = max(values or [1])
        buffered = max_value * 1.15
        if buffered <= 10:
            step = 1
        elif buffered <= 50:
            step = 5
        elif buffered <= 100:
            step = 10
        else:
            step = 20
        return max(step, int((buffered + step - 1) // step) * step)

    def find_row(title: str) -> int:
        for row in range(1, ws.max_row + 1):
            if ws.cell(row, 1).value == title:
                return row
        raise DataError(f"汇总表缺少区域: {title}")

    prod_title_row = find_row("生产事件单统计")
    prod_header_row = prod_title_row + 1
    prod_data_row = prod_title_row + 2
    type_title_row = find_row("告警类型统计")
    type_header_row = type_title_row + 1
    hf_title_row = find_row("高频告警事件")
    monitor_total_row = next(row for row in range(3, prod_title_row) if ws.cell(row, 1).value == "总计")

    bar = BarChart()
    bar.style = 10
    bar.title = "监控告警统计"
    bar.y_axis.title = "数量"
    data = Reference(ws, min_col=3, max_col=5, min_row=2, max_row=monitor_total_row)
    cats = Reference(ws, min_col=1, min_row=3, max_row=monitor_total_row)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    colors = ("4472C4", "ED7D31", "FFC000")
    for idx, series in enumerate(bar.series):
        series.graphicalProperties.solidFill = colors[idx]
        series.graphicalProperties.line.solidFill = colors[idx]
    bar.height = 8.5
    bar.width = 18
    ws.add_chart(bar, "G1")

    prod = PieChart()
    prod.style = 10
    prod.title = "事件单审结率"
    data = Reference(ws, min_col=2, max_col=3, min_row=prod_data_row, max_row=prod_data_row)
    cats = Reference(ws, min_col=2, max_col=3, min_row=prod_header_row, max_row=prod_header_row)
    prod.add_data(data, from_rows=True)
    prod.set_categories(cats)
    prod.dataLabels = DataLabelList()
    prod.dataLabels.showPercent = True
    prod.dataLabels.showCatName = True
    prod.dataLabels.showLeaderLines = True
    prod.firstSliceAng = 270
    prod.height = 8.5
    prod.width = 18
    ws.add_chart(prod, "G20")

    type_chart = BarChart()
    type_chart.type = "bar"
    type_chart.barDir = "bar"
    type_chart.style = 10
    type_chart.title = "告警类型统计"
    type_last_row = max(type_header_row + 1, hf_title_row - 2)
    type_values = [
        int(ws.cell(row, 2).value or 0)
        for row in range(type_header_row + 1, type_last_row + 1)
        if ws.cell(row, 1).value not in (None, "")
    ]
    data = Reference(ws, min_col=2, min_row=type_header_row, max_row=type_last_row)
    cats = Reference(ws, min_col=1, min_row=type_header_row + 1, max_row=type_last_row)
    type_chart.add_data(data, titles_from_data=True)
    type_chart.set_categories(cats)
    type_chart.legend = None
    type_chart.gapWidth = 45
    type_chart.x_axis.scaling.min = 0
    type_chart.x_axis.scaling.max = nice_axis_max(type_values)
    for series in type_chart.series:
        series.graphicalProperties.solidFill = "ED7D31"
        series.graphicalProperties.line.solidFill = "ED7D31"
        series.dLbls = DataLabelList()
        series.dLbls.showVal = True
    type_chart.height = min(8.5, max(5.8, 3.6 + 0.8 * max(len(type_values), 1)))
    type_chart.width = 13.5
    ws.add_chart(type_chart, "G39")


def setup_common_widths(ws, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def fill_block(ws, min_row: int, max_row: int, min_col: int, max_col: int, color: str) -> None:
    fill = PatternFill("solid", fgColor=color)
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            ws.cell(row, col).fill = fill


def border_block(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    thin = Side(style="thin", color="FF000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            ws.cell(row, col).border = border


def style_table(ws, header_row: int, first_data_row: int, last_row: int, last_col: int) -> None:
    thin = Side(style="thin", color="FF000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="FFFFFF00")
    for row in range(header_row, last_row + 1):
        for col in range(1, last_col + 1):
            cell = ws.cell(row, col)
            cell.border = border
            cell.alignment = Alignment(horizontal="center" if col == 1 else "left", vertical="center", wrap_text=True)
            if row == header_row:
                cell.fill = header_fill
                cell.font = Font(name="宋体", size=11, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif row >= first_data_row:
                cell.font = Font(name="宋体", size=11)


def write_summary(ws, data: dict[str, Any], start: date, end: date) -> None:
    for merged in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged))
    clear_values(ws, 1, max(ws.max_row, 60), 1, 7)

    alerts = all_alerts(data)
    days = sorted(
        {
            shift_date(dt)
            for alert in alerts
            for dt in [parse_dt(alert.get("通知时间"))]
            if dt
        }
    ) or [start]
    monitor_rows: list[list[Any]] = []
    for day in days:
        matched = [alert for alert in alerts if (dt := parse_dt(alert.get("通知时间"))) and shift_date(dt) == day]
        monitor_rows.append([
            f"{day:%m月%d日}",
            "",
            sum(1 for alert in matched if alert.get("告警来源") == "OMP平台"),
            sum(1 for alert in matched if "邮件电话" in str(alert.get("通知方式") or "")),
            sum(1 for alert in matched if alert.get("告警来源") == "青藤云"),
        ])

    summaries = data["qingteng_summary"] or [
        f"[青藤云安全][{row_value(row, '级别')}]{row_value(row, '告警信息')}" for row in data["qingteng_events"]
    ]
    summary_texts = list(dict.fromkeys(summaries))
    monitor_total_row = 3 + len(monitor_rows)
    qingteng_title_row = monitor_total_row + 2
    qingteng_first_row = qingteng_title_row + 1
    qingteng_last_row = qingteng_title_row + max(len(summary_texts), 1)
    prod_title_row = qingteng_last_row + 2
    prod_header_row = prod_title_row + 1
    prod_data_row = prod_title_row + 2
    os_title_row = prod_data_row + 1
    os_header_row = os_title_row + 1
    os_data_row = os_title_row + 2
    type_title_row = os_data_row + 2
    type_header_row = type_title_row + 1

    for rng in (
        "A1:E1",
        f"A{qingteng_title_row}:E{qingteng_title_row}",
        f"A{prod_title_row}:E{prod_title_row}",
        f"A{os_title_row}:E{os_title_row}",
        f"A{type_title_row}:C{type_title_row}",
    ):
        ws.merge_cells(rng)
    for row in range(qingteng_first_row, qingteng_last_row + 1):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)

    write_row(ws, 1, ["监控告警统计", "", "", "", ""])
    write_row(ws, 2, ["日期", "值班人", "监控告警数量", "邮件电话通知数量", "青藤云"])
    for idx, values in enumerate(monitor_rows, 3):
        write_row(ws, idx, values)
    totals = [sum(row[col] for row in monitor_rows) for col in (2, 3, 4)]
    write_row(ws, monitor_total_row, ["总计", "", *totals])

    write_row(ws, qingteng_title_row, ["青藤云安全告警汇总", "", "", "", ""])
    if summary_texts:
        for row_index, text in enumerate(summary_texts, qingteng_first_row):
            ws.cell(row_index, 1).value = text
            ws.cell(row_index, 1).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.row_dimensions[row_index].height = 34
    else:
        ws.cell(qingteng_first_row, 1).value = ""

    prod_rows = filter_production_events(data["production_events"], start, end)
    prod_total = len(prod_rows)
    prod_finished = sum(1 for row in prod_rows if row_value(row, "当前节点") == "审结" or parse_dt(row_value(row, "结束时间")))
    prod_processing = prod_total - prod_finished
    report_end = production_report_end(end)
    prod_timeout = sum(1 for row in prod_rows if is_production_timeout(row, report_end))
    write_row(ws, prod_title_row, ["生产事件单统计", "", "", "", ""])
    write_row(ws, prod_header_row, ["总数", "审结", "处理中/事件录入", "超时", "审结率"])
    write_row(ws, prod_data_row, [prod_total, prod_finished, prod_processing, prod_timeout, percent(prod_finished, prod_total)])

    write_row(ws, os_title_row, ["操作系统版本升级事件单统计", "", "", "", ""])
    write_row(ws, os_header_row, ["总数", "审结", "安全团队待处理", "负责人待审核", "处理中"])
    write_row(ws, os_data_row, [0, 0, 0, 0, 0])

    type_counter: Counter[str] = Counter()
    for alert in alerts:
        key = alert.get("告警指标名称") if alert.get("告警来源") == "OMP平台" else alert.get("事件名称")
        type_counter[str(key or "其他").strip()] += 1
    total_alert_events = sum(type_counter.values())
    alert_types = [(name, count) for name, count in type_counter.most_common() if count >= 5]

    write_row(ws, type_title_row, ["告警类型统计", "", ""])
    write_row(ws, type_header_row, ["事件类型", "出现次数", "占比"])
    for row_index, (name, count) in enumerate(alert_types, type_header_row + 1):
        write_row(ws, row_index, [name, count, percent(count, total_alert_events)])

    hf_counter: Counter[tuple[str, str, str]] = Counter()
    for alert in alerts:
        detail = str(alert.get("告警信息") or "").strip()
        if detail:
            hf_counter[(detail, str(alert.get("所属应用") or ""), str(alert.get("负责人") or ""))] += 1
    hf_rows = [(detail, count, app, owner) for (detail, app, owner), count in hf_counter.most_common() if count >= 5]
    hf_start = max(type_header_row + 3 + len(alert_types), type_title_row + 8)
    ws.merge_cells(start_row=hf_start, start_column=1, end_row=hf_start, end_column=4)
    write_row(ws, hf_start, ["高频告警事件", "", "", ""])
    write_row(ws, hf_start + 1, ["告警详情", "出现次数", "应用", "管理员"])
    for row_index, values in enumerate(hf_rows, hf_start + 2):
        write_row(ws, row_index, list(values))

    login_row = hf_start + 3 + len(hf_rows)
    ws.merge_cells(start_row=login_row, start_column=1, end_row=login_row, end_column=5)
    ws.cell(login_row, 1).value = f"青藤云异常登录事件: {data['qingteng_abnormal_login_count']}"

    for row in range(1, login_row + 1):
        for col in range(1, 8):
            cell = ws.cell(row, col)
            cell.font = Font(name="宋体", size=11, bold=row in (1, 2, qingteng_title_row, prod_title_row, prod_header_row, os_title_row, os_header_row, type_title_row, type_header_row, hf_start, hf_start + 1))
            cell.alignment = Alignment(horizontal="center" if col <= 5 else "left", vertical="center", wrap_text=True)
    fill_block(ws, 1, monitor_total_row, 1, 5, "FFE2F0D9")
    fill_block(ws, qingteng_title_row, qingteng_last_row, 1, 5, "FFFFC000")
    fill_block(ws, prod_title_row, os_data_row, 1, 5, "FFD9E2F3")
    fill_block(ws, type_title_row, type_header_row + max(len(alert_types), 1), 1, 3, "FFFCE4D6")
    fill_block(ws, hf_start, max(hf_start + 1, hf_start + 1 + len(hf_rows)), 1, 4, "FFFFF2CC")
    fill_block(ws, login_row, login_row, 1, 5, "FFE2F0D9")
    fill_block(ws, prod_header_row, prod_data_row, 4, 4, "FFFFFF00")
    ws.cell(prod_header_row, 4).font = Font(name="宋体", size=12, bold=True)
    ws.cell(prod_data_row, 4).font = Font(name="宋体", size=12, bold=True)
    for row in (1, qingteng_title_row, prod_title_row, os_title_row, type_title_row, hf_start):
        for col in range(1, 6):
            ws.cell(row, col).font = Font(name="宋体", size=14, bold=True)
    for row in (2, monitor_total_row, prod_header_row, prod_data_row, os_header_row, os_data_row, type_header_row, hf_start + 1):
        for col in range(1, 6):
            ws.cell(row, col).font = Font(name="宋体", size=12, bold=True if row in (2, monitor_total_row, prod_header_row, os_header_row, type_header_row, hf_start + 1) else False)
    border_block(ws, 1, monitor_total_row, 1, 5)
    border_block(ws, qingteng_title_row, qingteng_last_row, 1, 5)
    border_block(ws, prod_title_row, os_data_row, 1, 5)
    border_block(ws, type_title_row, type_header_row + max(len(alert_types), 1), 1, 3)
    border_block(ws, hf_start, max(hf_start + 1, hf_start + 1 + len(hf_rows)), 1, 4)
    border_block(ws, login_row, login_row, 1, 5)
    for row_index in range(qingteng_first_row, qingteng_last_row + 1):
        ws.cell(row_index, 1).font = Font(name="宋体", size=11, bold=True, color="FF000000")
        ws.cell(row_index, 1).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.cell(login_row, 1).font = Font(name="宋体", size=14, bold=True)
    ws.cell(login_row, 1).alignment = Alignment(horizontal="center", vertical="center")
    setup_common_widths(ws, {"A": 24, "B": 16, "C": 20, "D": 20, "E": 14, "G": 16, "H": 16, "I": 16, "J": 16, "K": 16, "L": 16})
    ensure_summary_charts(ws)


def write_production_sheet(ws, rows: list[dict[str, Any]], start: date, end: date) -> None:
    for merged in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged))
    reset_sheet_rows(ws, 2)
    ws.merge_cells("A1:H1")
    ws["A1"] = "生产事件工单"
    ws["A1"].font = Font(name="宋体", size=16, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    for col, header in enumerate(PROD_HEADERS, 1):
        ws.cell(2, col).value = header
    report_end = production_report_end(end)
    prod_rows = filter_production_events(rows, start, end)
    for row_index, row in enumerate(prod_rows, 3):
        values = [fmt_dt(row_value(row, header)) if "时间" in header else row_value(row, header) for header in PROD_HEADERS]
        write_row(ws, row_index, values)
    style_table(ws, 2, 3, max(2, ws.max_row), len(PROD_HEADERS))
    timeout_fill = PatternFill("solid", fgColor="FFFFC7CE")
    timeout_font = Font(name="宋体", size=11, color="FF9C0006")
    finished_fill = PatternFill("solid", fgColor="FFC6EFCE")
    normal_fill = PatternFill("solid", fgColor="FFFFFFFF")
    for row_index, row in enumerate(prod_rows, 3):
        if is_production_timeout(row, report_end):
            fill = timeout_fill
            font = timeout_font
        elif is_finished(row):
            fill = finished_fill
            font = Font(name="宋体", size=11)
        else:
            fill = normal_fill
            font = Font(name="宋体", size=11)
        for col in range(1, len(PROD_HEADERS) + 1):
            ws.cell(row_index, col).fill = fill
            ws.cell(row_index, col).font = font
    setup_common_widths(ws, {"A": 18, "B": 12, "C": 80, "D": 16, "E": 22, "F": 22, "G": 22, "H": 22})


def write_os_sheet(ws, rows: list[dict[str, Any]]) -> None:
    reset_sheet_rows(ws, 1)
    for col, header in enumerate(OS_HEADERS, 1):
        ws.cell(1, col).value = header
    sorted_rows = sort_os_baseline_events(rows)
    for row_index, row in enumerate(sorted_rows, 2):
        values = [fmt_dt(row_value(row, header)) if "时间" in header else row_value(row, header) for header in OS_HEADERS]
        write_row(ws, row_index, values)
    style_table(ws, 1, 2, max(1, ws.max_row), len(OS_HEADERS))
    finished_fill = PatternFill("solid", fgColor="FFC6EFCE")
    normal_fill = PatternFill("solid", fgColor="FFFFFFFF")
    for row_index, row in enumerate(sorted_rows, 2):
        fill = finished_fill if is_finished(row) else normal_fill
        for col in range(1, len(OS_HEADERS) + 1):
            ws.cell(row_index, col).fill = fill
    setup_common_widths(ws, {"A": 22, "B": 12, "C": 60, "D": 24, "E": 22, "F": 22, "G": 22, "H": 22, "I": 12, "J": 18})


def write_empty_os_version_sheet(ws) -> None:
    reset_sheet_rows(ws, 1)
    for col, header in enumerate(OS_HEADERS, 1):
        ws.cell(1, col).value = header
    style_table(ws, 1, 2, 1, len(OS_HEADERS))
    setup_common_widths(ws, {"A": 22, "B": 12, "C": 60, "D": 24, "E": 22, "F": 22, "G": 22, "H": 22, "I": 12, "J": 18})


def get_or_create_day_sheet(wb, day: date):
    title = f"{day:%m月%d日}"
    if title in wb.sheetnames:
        ws = wb[title]
        reset_sheet_rows(ws, 0)
    else:
        ws = wb.create_sheet(title)
    return ws


def delete_unused_day_sheets(wb, days: list[date]) -> None:
    keep = {f"{day:%m月%d日}" for day in days}
    for name in list(wb.sheetnames):
        if re.fullmatch(r"\d{2}月\d{2}日", name) and name not in keep:
            del wb[name]


def write_day_sheet(wb, day: date, alerts: list[dict[str, Any]]) -> None:
    ws = get_or_create_day_sheet(wb, day)
    for col, header in enumerate(ALERT_HEADERS, 1):
        ws.cell(1, col).value = header

    day_alerts = [alert for alert in alerts if in_day_window(parse_dt(alert.get("通知时间")), day)]
    day_alerts.sort(key=lambda alert: (0 if alert.get("告警来源") == "OMP平台" else 1, parse_dt(alert.get("通知时间")) or datetime.min))

    for row_index, alert in enumerate(day_alerts, 2):
        write_row(
            ws,
            row_index,
            [row_index - 1, alert.get("IP地址", ""), alert.get("告警来源", ""), alert.get("级别", ""), alert.get("所属应用", ""), alert.get("负责人", ""), alert.get("告警信息", "")],
        )

    summary_start = len(day_alerts) + 3
    by_source = Counter(str(alert.get("告警来源") or "") for alert in day_alerts)
    ws.cell(summary_start, 7).value = f"{day.month}月{day.day}日总结:"
    for offset, (source, count) in enumerate(sorted(by_source.items()), 1):
        ws.cell(summary_start + offset, 7).value = f"{source}告警处理{count}条"

    style_table(ws, 1, 2, max(1, ws.max_row), len(ALERT_HEADERS))
    for row in range(summary_start, ws.max_row + 1):
        for col in range(1, len(ALERT_HEADERS) + 1):
            ws.cell(row, col).border = Border()
            ws.cell(row, col).fill = PatternFill(fill_type=None)
    setup_common_widths(ws, {"A": 8, "B": 18, "C": 16, "D": 12, "E": 28, "F": 28, "G": 120})


def build_report_from_template(
    data: dict[str, Any],
    template_path: str,
    output_path: str,
    start: date | None = None,
    end: date | None = None,
) -> tuple[date, date]:
    template = Path(template_path)
    output = Path(output_path)
    if not template.exists():
        raise WorkbookError(f"Template file not found: {template}")

    if start and end is None:
        end = start + timedelta(days=6)
    elif start is None or end is None:
        inferred_start, inferred_end = infer_range(data)
        start = start or inferred_start
        end = end or inferred_end
    if start > end:
        raise DataError("start_date cannot be later than end_date")

    output.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template, output)
    wb = load_workbook(output)

    for sheet_name in (SUMMARY_SHEET, PROD_SHEET, OS_VERSION_SHEET, OS_BASELINE_SHEET):
        if sheet_name not in wb.sheetnames:
            raise WorkbookError(f"Template is missing required sheet: {sheet_name}")

    write_summary(wb[SUMMARY_SHEET], data, start, end)
    write_production_sheet(wb[PROD_SHEET], data["production_events"], start, end)
    write_empty_os_version_sheet(wb[OS_VERSION_SHEET])
    write_os_sheet(wb[OS_BASELINE_SHEET], data["os_baseline_events"])

    days = alert_days(data) or iter_days(start, end)
    delete_unused_day_sheets(wb, days)
    alerts = all_alerts(data)
    for day in days:
        write_day_sheet(wb, day, alerts)

    order = [SUMMARY_SHEET, PROD_SHEET, OS_VERSION_SHEET, OS_BASELINE_SHEET] + [f"{day:%m月%d日}" for day in days]
    wb._sheets = [wb[name] for name in order if name in wb.sheetnames] + [ws for ws in wb.worksheets if ws.title not in order]
    wb.save(output)
    wb.close()
    return start, end


def generate_ecc_report_from_template(
    template_file: str = DEFAULT_TEMPLATE,
    output_file: str = "",
    production_events: Any = "",
    os_baseline_events: Any = "",
    omp_alerts: Any = "",
    qingteng_events: Any = "",
    qingteng_summary: Any = "",
    qingteng_abnormal_login_count: Any = "",
    start_date: str = "",
    end_date: str = "",
) -> dict[str, Any]:
    if not output_file:
        output_file = f"reports/ECC_{datetime.now():%Y%m%d_%H%M%S}.xlsx"

    data = make_source_data(
        production_events=production_events,
        os_baseline_events=os_baseline_events,
        omp_alerts=omp_alerts,
        qingteng_events=qingteng_events,
        qingteng_summary=qingteng_summary,
        qingteng_abnormal_login_count=qingteng_abnormal_login_count,
    )
    used_start, used_end = build_report_from_template(
        data=data,
        template_path=template_file,
        output_path=output_file,
        start=parse_date_arg(start_date),
        end=parse_date_arg(end_date),
    )
    return {
        "success": True,
        "message": f"已生成: {output_file}",
        "file_path": output_file,
        "template_file": template_file,
        "start_date": f"{used_start:%Y-%m-%d}",
        "end_date": f"{used_end:%Y-%m-%d}",
    }


def generate_ecc_report_from_template_files(
    template_file: str = DEFAULT_TEMPLATE,
    output_file: str = "",
    production_events_file: str = "",
    os_baseline_events_file: str = "",
    omp_alerts_file: str = "",
    qingteng_events_file: str = "",
    qingteng_summary_file: str = "",
    qingteng_abnormal_login_count_file: str = "",
    start_date: str = "",
    end_date: str = "",
) -> dict[str, Any]:
    if not output_file:
        output_file = f"reports/ECC_{datetime.now():%Y%m%d_%H%M%S}.xlsx"

    abnormal_login_payload = load_json_file(qingteng_abnormal_login_count_file) if qingteng_abnormal_login_count_file else 0
    if isinstance(abnormal_login_payload, dict):
        abnormal_login_payload = abnormal_login_payload.get("count", abnormal_login_payload.get("value", 0))

    data = make_source_data(
        production_events=load_json_file(production_events_file) if production_events_file else [],
        os_baseline_events=load_json_file(os_baseline_events_file) if os_baseline_events_file else [],
        omp_alerts=load_json_file(omp_alerts_file) if omp_alerts_file else [],
        qingteng_events=load_json_file(qingteng_events_file) if qingteng_events_file else [],
        qingteng_summary=load_json_file(qingteng_summary_file) if qingteng_summary_file else [],
        qingteng_abnormal_login_count=abnormal_login_payload,
    )
    used_start, used_end = build_report_from_template(
        data=data,
        template_path=template_file,
        output_path=output_file,
        start=parse_date_arg(start_date),
        end=parse_date_arg(end_date),
    )
    return {
        "success": True,
        "message": f"已生成: {output_file}",
        "file_path": output_file,
        "template_file": template_file,
        "start_date": f"{used_start:%Y-%m-%d}",
        "end_date": f"{used_end:%Y-%m-%d}",
    }
