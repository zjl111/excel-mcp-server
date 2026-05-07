import os
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

_REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_SRC = os.path.join(_REPO_ROOT, "src")
if _SRC not in sys.path:
    sys.path.insert(0, _SRC)

from excel_mcp.ecc_report import (  # noqa: E402
    generate_ecc_report_from_template,
    generate_ecc_report_from_template_files,
)


class TestEccTemplateReport(unittest.TestCase):
    def test_generate_report_from_template(self):
        template = Path(_REPO_ROOT) / "excel_files" / "templates" / "ECC_TEMPLATE.xlsx"
        self.assertTrue(template.exists(), "ECC template must exist for report generation")

        with tempfile.TemporaryDirectory() as tmpdir:
            output = Path(tmpdir) / "ECC_test.xlsx"
            result = generate_ecc_report_from_template(
                template_file=str(template),
                output_file=str(output),
                production_events=[{
                    "工单编号": "P1",
                    "创建人": "何红",
                    "流程描述": "问题一",
                    "当前节点": "问题处理",
                    "当前处理人/角色": "何红",
                    "送达时间": "2026-04-15 08:53:11",
                    "创建时间": "2026-04-15 08:49:14",
                    "结束时间": "",
                }],
                os_baseline_events=[{
                    "工单编号": "B1",
                    "创建人": "监控岗王越",
                    "流程描述": "基线升级",
                    "当前节点": "审结",
                    "当前处理人/角色": "",
                    "送达时间": "2026-02-25 14:03:13",
                    "创建时间": "2026-02-25 14:03:13",
                    "结束时间": "2026-02-26 10:00:00",
                    "是否启阅": "已阅",
                    "主机": "172.22.2.106",
                }],
                omp_alerts=[{
                    "IP 地址": "172.16.130.46",
                    "告警来源": "OMP 平台",
                    "级别": "次要",
                    "所属应用": "恒生投资系统 03.2",
                    "负责人": "何红",
                    "告警信息": "CPU使用率",
                    "告警指标名称": "CPU使用率",
                    "通知时间": "2026-04-15 23:56:27",
                    "通知方式": "邮件电话",
                }],
                qingteng_events=[{
                    "IP 地址": "172.16.130.37",
                    "告警来源": "青藤云",
                    "级别": "信息",
                    "所属应用": "恒生投资系统 03.2",
                    "负责人": "",
                    "告警信息": "WinWeb 命令执行",
                    "事件名称": "[WinWeb 命令执行]",
                    "通知时间": "2026-04-16 20:52:05",
                    "通知方式": "微信",
                }],
                qingteng_summary=["[青藤云安全][信息] [TZJY240531W06R/172.16.130.37] WinWeb 命令执行"],
                qingteng_abnormal_login_count=100,
            )
            self.assertTrue(result["success"])
            self.assertTrue(output.exists())

            wb = load_workbook(output)
            self.assertEqual(
                wb.sheetnames[:4],
                ["汇总", "生产事件单处理情况", "操作系统版本升级事件单处理情况", "操作系统基线要求升级事件单处理情况"],
            )
            self.assertIn("04月15日", wb.sheetnames)
            self.assertIn("04月16日", wb.sheetnames)
            self.assertEqual(len(wb["汇总"]._charts), 3)
            self.assertEqual(wb["生产事件单处理情况"]["A3"].value, "P1")
            self.assertEqual(wb["生产事件单处理情况"]["A3"].fill.fgColor.rgb, "FFFFFFFF")
            self.assertEqual(wb["操作系统基线要求升级事件单处理情况"]["A2"].value, "B1")
            self.assertEqual(wb["操作系统基线要求升级事件单处理情况"]["A2"].fill.fgColor.rgb, "FFC6EFCE")
            wb.close()

    def test_generate_report_from_template_files(self):
        template = Path(_REPO_ROOT) / "excel_files" / "templates" / "ECC_TEMPLATE.xlsx"
        self.assertTrue(template.exists(), "ECC template must exist for report generation")

        with tempfile.TemporaryDirectory() as tmpdir:
            tmpdir_path = Path(tmpdir)
            (tmpdir_path / "production_events.json").write_text(
                '[{"工单编号":"P2","创建人":"何红","流程描述":"问题二","当前节点":"问题处理","当前处理人/角色":"何红","送达时间":"2026-04-15 08:53:11","创建时间":"2026-04-15 08:49:14","结束时间":""}]',
                encoding="utf-8",
            )
            (tmpdir_path / "os_baseline_events.json").write_text(
                '[{"工单编号":"B2","创建人":"监控岗王越","流程描述":"基线升级","当前节点":"处理中","当前处理人/角色":"孙赫","送达时间":"2026-02-25 14:03:13","创建时间":"2026-02-25 14:03:13","结束时间":"","是否启阅":"已阅","主机":"172.22.2.106"}]',
                encoding="utf-8",
            )
            (tmpdir_path / "omp_alerts.json").write_text(
                '[{"IP 地址":"172.16.130.46","告警来源":"OMP 平台","级别":"次要","所属应用":"恒生投资系统 03.2","负责人":"何红","告警信息":"CPU使用率","告警指标名称":"CPU使用率","通知时间":"2026-04-15 23:56:27","通知方式":"邮件电话"}]',
                encoding="utf-8",
            )
            (tmpdir_path / "qingteng_events.json").write_text(
                '[{"IP 地址":"172.16.130.37","告警来源":"青藤云","级别":"信息","所属应用":"恒生投资系统 03.2","负责人":"","告警信息":"WinWeb 命令执行","事件名称":"[WinWeb 命令执行]","通知时间":"2026-04-16 20:52:05","通知方式":"微信"}]',
                encoding="utf-8",
            )
            (tmpdir_path / "qingteng_summary.json").write_text(
                '["[青藤云安全][信息] [TZJY240531W06R/172.16.130.37] WinWeb 命令执行"]',
                encoding="utf-8",
            )
            (tmpdir_path / "qingteng_abnormal_login_count.json").write_text(
                '{"count": 100}',
                encoding="utf-8",
            )

            output = tmpdir_path / "ECC_files_test.xlsx"
            result = generate_ecc_report_from_template_files(
                template_file=str(template),
                output_file=str(output),
                production_events_file=str(tmpdir_path / "production_events.json"),
                os_baseline_events_file=str(tmpdir_path / "os_baseline_events.json"),
                omp_alerts_file=str(tmpdir_path / "omp_alerts.json"),
                qingteng_events_file=str(tmpdir_path / "qingteng_events.json"),
                qingteng_summary_file=str(tmpdir_path / "qingteng_summary.json"),
                qingteng_abnormal_login_count_file=str(tmpdir_path / "qingteng_abnormal_login_count.json"),
            )
            self.assertTrue(result["success"])
            self.assertTrue(output.exists())
            wb = load_workbook(output)
            self.assertEqual(wb["生产事件单处理情况"]["A3"].value, "P2")
            self.assertEqual(wb["操作系统基线要求升级事件单处理情况"]["A2"].value, "B2")
            self.assertEqual(wb["操作系统基线要求升级事件单处理情况"]["A2"].fill.fgColor.rgb, "FFFFFFFF")
            wb.close()


if __name__ == "__main__":
    unittest.main()
